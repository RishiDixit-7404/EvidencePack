"""Tests for the hardened demo controls workbook."""

from pathlib import Path

from openpyxl import load_workbook


EXPECTED_COLUMNS = [
    "Control ID",
    "Control Description",
    "Required Evidence Type",
    "Domain",
]
EXPECTED_DOMAINS = {
    "Access Management",
    "Change Management",
    "Data Protection",
}
EXPECTED_CONTROL_IDS = {
    "AM-01",
    "AM-02",
    "AM-03",
    "AM-04",
    "CM-01",
    "CM-02",
    "CM-03",
    "CM-04",
    "DP-01",
    "DP-02",
    "DP-03",
    "DP-04",
}


def _load_controls_rows() -> list[tuple[str, str, str, str]]:
    workbook_path = Path("demo_evidence") / "controls.xlsx"
    assert workbook_path.exists()

    workbook = load_workbook(workbook_path)
    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    assert rows
    assert list(rows[0]) == EXPECTED_COLUMNS
    return rows[1:]


def test_demo_controls_workbook_schema_and_content() -> None:
    rows = _load_controls_rows()

    assert len(rows) == 12
    assert {row[3] for row in rows} == EXPECTED_DOMAINS
    assert {row[0] for row in rows} == EXPECTED_CONTROL_IDS


def test_demo_controls_required_cells_are_complete() -> None:
    rows = _load_controls_rows()
    control_ids = [row[0] for row in rows]

    assert len(control_ids) == len(set(control_ids))
    for row in rows:
        assert all(str(value).strip() for value in row)
        assert str(row[2]).strip()

